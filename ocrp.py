import os
from llama_index.llms.openai import OpenAI
from llama_index.core.node_parser import SimpleNodeParser
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core import (
    SimpleDirectoryReader,
    load_index_from_storage,
    VectorStoreIndex,
    StorageContext,
    ServiceContext
)
from llama_index.vector_stores.faiss import FaissVectorStore
from IPython.display import Markdown, display
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from llama_index.core import get_response_synthesizer 
from llama_index.core.retrievers import (
    BaseRetriever,
    VectorIndexRetriever,
    KeywordTableSimpleRetriever,
)
from llama_index.core.query_engine import RetrieverQueryEngine
import re
import ast
from openpyxl import load_workbook






OPENAI_API_KEY=''


def process_pdf(pdf_path, query, output_path):
    
    print(f"Processing {pdf_path}")
    loader = SimpleDirectoryReader(input_files=[pdf_path]
            )
    documents = loader.load_data()

    llm = OpenAI(model="gpt-4o", temperature=0)
    service_context = ServiceContext.from_defaults(llm=llm)

    
    index = VectorStoreIndex.from_documents(documents, service_context=service_context)

    retriever = VectorIndexRetriever(
        index=index,
        similarity_top_k=15 
    )

    query_engine = RetrieverQueryEngine(
        retriever=retriever  
    )

    response = query_engine.query(query)
    
    return response
    

def save_response_to_txt(response,pdf_name, output_path):
    if response is None:
        print("No response to process.")
        return
    
    response_data = response.response
    
    try:
        with open(output_path, 'a') as file:  
            file.write(f"PDF Name: {pdf_name}\n")
            file.write(response_data + '\n')  
        print(f"Response data appended to {output_path}")
    except Exception as e:
        print(f"Error occurred while appending to {output_path}: {str(e)}")



def save_to_excel(response, pdf_name):
    try:
        # Extract JSON data from response
        json_string = str(response.response)
        a = json_string.replace('json\n', '', 1).strip('```')
        data = json.loads(a)
        
        # Create DataFrame
        df = pd.DataFrame(data)

        # Add PDF name and reorder columns
        df["Nome del PDF"] = pdf_name
        columns_order = ["Nome del PDF"] + [col for col in df.columns if col != "Nome del PDF"]
        df = df[columns_order]

        excel_file = 'output.xlsx'

        if os.path.exists(excel_file):
            # Read existing Excel file
            book = load_workbook(excel_file)
            sheet = book.active
            
            # Find the last row with data
            last_row = sheet.max_row
            
            # Skip a row before writing new data
            start_row = last_row + 2
            
            # Write headers
            for c_idx, header in enumerate(df.columns, start=1):
                sheet.cell(row=start_row, column=c_idx, value=header)
            
            # Write new data starting from the next row
            for r_idx, row in enumerate(df.itertuples(index=False), start=start_row+1):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Save the workbook
            book.save(excel_file)
        else:
            # Create a new Excel file
            df.to_excel(excel_file, index=False)

        print(f"Data has been appended to {excel_file}")
        return "Data saved successfully"
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return f"Error saving data: {str(e)}"


def extract_info_from_pdf(directory_path, query,output_path):
    count = 0
    for filename in os.listdir(directory_path):
        if filename.endswith('.pdf'):
            pdf_name = filename
            pdf_path = os.path.join(directory_path, filename)
            print("pdf path", pdf_path)
            response = process_pdf(pdf_path, query, output_path)
            save_to_excel(response,pdf_name)
            count += 1
            if count >= 6:
                break
            


if __name__ == "__main__":
    pdf_path = "/home/me/software/pdfocr/schede/"
    query = """
    Cerca la tabella intitolata SEZIONE 3 nel documento.
    Estrai tutti i dati dalla tabella, incluse le intestazioni delle colonne.
    Formatta i dati come una stringa JSON, dove ogni riga della tabella è un oggetto in un array.
    Usa le intestazioni delle colonne come chiavi per ogni oggetto.
    Se non riesci a formattare come JSON, restituisci i dati della tabella in un formato tabulare usando il carattere '|' come separatore.
    """
    excel_output_path = "output_table.txt"

    response = extract_info_from_pdf(pdf_path, query,excel_output_path)
    